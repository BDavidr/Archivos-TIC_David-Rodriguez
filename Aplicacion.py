import serial
import time
import sys
from PyQt5.QtWidgets import QDialog, QApplication, QTableWidgetItem
from FrmAplicacion import *
from PyQt5.QtWidgets import QDialog, QApplication, QTableWidgetItem
from FrmAplicacion import *
import math
import pandas as pd
import openpyxl
from tkinter import Tk, filedialog
from tkinter.filedialog import asksaveasfilename

from PyQt5.QtCore import QTimer,Qt


class mostrar(QDialog):
    def __init__(self):
        super().__init__()

        self.primeraventana = Ui_FrmAplicacion2()
        self.primeraventana.setupUi(self)

        # Progess Bar
        self.timerRecepcion = QTimer(self)
        self.timerRecepcion.timeout.connect(self.actualizarContadorRecepcion)
        self.timerDescarga = QTimer(self)
        self.timerDescarga.timeout.connect(self.actualizarContadorDescarga)
        self.primeraventana.progressBarRecepcion.setValue(0)
        self.primeraventana.progressBarDescarga.setValue(0)
        #self.timer = QTimer(self)
        #self.timer.timeout.connect(self.actualizar_progreso)
        # self.pushButton.clicked.connect(self.iniciar_detener_progreso)
        # self.lineEdit.textChanged.connect(self.actualizar_intervalo)
        # Eventos click
        self.primeraventana.btnConectar.clicked.connect(self.ejecutarConexion)
        self.primeraventana.btnTrasmision.clicked.connect(self.EnviarOrdenes)
        self.primeraventana.btnConsulta.clicked.connect(self.ObtenerDatos)
        self.primeraventana.btnGenerarExel.clicked.connect(self.GenerarArchivoEXEL)

        self.show()
        # Variables y listas Globales
        self.puertoSerial = ''  # Puerto serial, almacena el puerto COM
        self.conectar = serial.Serial()  # Conexion serial, almacena la conectar serial
        self.DatosCompletosGlobales = b''
        # Funciones Iniciales
        self.desaparecerComponentes()
        # self.reaparecerComponentesRx()
        # self.reaparecerComponentesTx()

    def desaparecerComponentes(self):
        self.primeraventana.btnTrasmision.setEnabled(False)
        self.primeraventana.progressBarRecepcion.setEnabled(False)
        self.primeraventana.spinBoxNumeroTransmisiones.setEnabled(False)
        self.primeraventana.spinBoxTiempoEntreTramas.setEnabled(False)
        self.primeraventana.spinBoxNumeroBloques.setEnabled(False)
        self.primeraventana.spinBoxTiempoEntreBloques.setEnabled(False)
        self.primeraventana.spinBoxPotenciaTransmision.setEnabled(False)
        self.primeraventana.btnConsulta.setEnabled(False)
        self.primeraventana.progressBarDescarga.setEnabled(False)
        self.primeraventana.btnGenerarExel.setEnabled(False)
        self.primeraventana.label_DescargandoTramas.hide()
        self.primeraventana.label_EnviandoTramas.hide()

    def reaparecerComponentesPrimerEnvio(self):
        self.primeraventana.btnTrasmision.setEnabled(True)
        self.primeraventana.progressBarRecepcion.setEnabled(True)
        self.primeraventana.spinBoxNumeroTransmisiones.setEnabled(True)
        self.primeraventana.spinBoxTiempoEntreTramas.setEnabled(True)
        self.primeraventana.spinBoxNumeroBloques.setEnabled(True)
        self.primeraventana.spinBoxTiempoEntreBloques.setEnabled(True)
        self.primeraventana.spinBoxPotenciaTransmision.setEnabled(True)
        self.primeraventana.btnConsulta.setEnabled(True)
        self.primeraventana.progressBarDescarga.setEnabled(True)
        self.primeraventana.btnGenerarExel.setEnabled(True)

    def ejecutarConexion(self):  # Configuracion de conectar serial
        try:
            SeleccionDePuerto = self.primeraventana.comboxPuertos.itemText(
                self.primeraventana.comboxPuertos.currentIndex())
            self.puertoSerial = SeleccionDePuerto[0:3] + SeleccionDePuerto[4]
            self.conectar = serial.Serial(self.puertoSerial, 9600, 8, 'N', stopbits=1, timeout=None)
            time.sleep(1)
            self.conectar.close()
            print('Conexión Exitosa')
            self.reaparecerComponentesPrimerEnvio()
        except:
            print('Conexión Fallida !')

    def EnviarOrdenes(self):

        self.conectar.open()
        self.conectar.flush()
        sparador = b'\t'
        identificador = b'\n'
        msj = 'ini'  # Inicio de Ordenes
        mensajeCod = msj.encode()
        self.conectar.write(mensajeCod + sparador)
        conf = self.conectar.read_until(b'\xFF')
        print('Confirmacion: ' + conf[:5].decode())
        # Orden 'a' Número de Transmisiones
        msj = self.primeraventana.spinBoxNumeroTransmisiones.text() + 'a'  # Orden 'a' Numero de Transmiciones
        NumeroTX = int(self.primeraventana.spinBoxNumeroTransmisiones.text())
        mensajeCod = msj.encode()
        self.conectar.write(mensajeCod + sparador)
        conf = self.conectar.read_until(b'\xFF')
        print('Confirmacion: ' + conf[:5].decode())
        # Orden 'b' Tiempo entre tramas
        msj = self.primeraventana.spinBoxTiempoEntreTramas.text() + 'b'
        TiempoEntreTr = int(self.primeraventana.spinBoxTiempoEntreTramas.text())
        mensajeCod = msj.encode()
        self.conectar.write(mensajeCod + sparador)
        conf = self.conectar.read_until(b'\xFF')
        print('Confirmacion: ' + conf[:5].decode())
        # Orden 'c' Número de Bloques
        msj = self.primeraventana.spinBoxNumeroBloques.text() + 'c'
        NumeroBl = int(self.primeraventana.spinBoxNumeroBloques.text())
        mensajeCod = msj.encode()
        self.conectar.write(mensajeCod + sparador)
        conf = self.conectar.read_until(b'\xFF')
        print('Confirmacion: ' + conf[:5].decode())
        # Orden 'd' Tiempo entre Bloques
        msj = self.primeraventana.spinBoxTiempoEntreBloques.text() + 'd'
        TiempoEntreBl = int(self.primeraventana.spinBoxTiempoEntreBloques.text())
        mensajeCod = msj.encode()
        self.conectar.write(mensajeCod + sparador)
        conf = self.conectar.read_until(b'\xFF')
        print('Confirmacion: ' + conf[:5].decode())
        # Orden 'e' Potencia de transmisión
        msj = self.primeraventana.spinBoxPotenciaTransmision.text() + 'e'
        mensajeCod = msj.encode()
        self.conectar.write(mensajeCod + sparador)
        conf = self.conectar.read_until(b'\xFF')
        print('Confirmacion: ' + conf[:5].decode())
        # Orden 'f' Potencia de transmisión
        tamañoTrama = self.primeraventana.comboxTamanioTrama.itemText(
            self.primeraventana.comboxTamanioTrama.currentIndex())
        ordenTamañoTrama = '0'
        if tamañoTrama == 'Trama Corta':
            ordenTamañoTrama = '1'
        elif tamañoTrama == 'Trama Larga':
            ordenTamañoTrama = '2'
        msj = ordenTamañoTrama + 'f'
        mensajeCod = msj.encode()
        self.conectar.write(mensajeCod + sparador)
        conf = self.conectar.read_until(b'\xFF')
        print('Confirmacion: ' + conf[:5].decode())
        # Finalización de órdenes
        msj = 'fin'
        mensajeCod = msj.encode()
        self.conectar.write(mensajeCod + sparador)
        conf = self.conectar.read_until(b'\xFF')
        print('Confirmacion: ' + conf[:5].decode())
        self.conectar.close()  # Cerrar conexión
        print('Ordenes Completas ')
        #########################################################################
        #Configuración de tiempos
        tiempoAdicionalBl = (TiempoEntreBl) * 0.55
        tiempoAdicionalTr = (TiempoEntreTr) * 0.55
        self.tiempoTotal=0.0
        if NumeroBl==1:
            self.tiempoTotal = (TiempoEntreTr*NumeroTX) / 1000
        elif NumeroBl>1:
            self.tiempoTotal = (((TiempoEntreTr+tiempoAdicionalTr) * (NumeroTX)* (NumeroBl)) / 1000) + ((TiempoEntreBl+ tiempoAdicionalBl) * (NumeroBl-1))
        print(f"tiempoTotal: {self.tiempoTotal}")
        self.iniciarTemporizadorRecepcion()

    def iniciarTemporizadorRecepcion(self):
        self.total_pasos = 100
        self.contador = 0
        tiempo_en_segundos = float(self.tiempoTotal)
        self.primeraventana.progressBarRecepcion.setValue(0)
        # Establece el intervalo del temporizador para avanzar la barra de progreso en el tiempo ingresado
        self.timer_intervalo = int(tiempo_en_segundos * 1000 / self.total_pasos)
        self.timerRecepcion.start(self.timer_intervalo)

    def actualizarContadorRecepcion(self):
        self.contador += 1
        self.primeraventana.progressBarRecepcion.setValue(self.contador)# Actualiza la barra de progreso
        if self.contador == self.total_pasos:
            self.timerRecepcion.stop()

    def ObtenerDatos(self):
        self.conectar.open() # Abrir el puerto serial
        self.conectar.flush()
        sparador = b'\r'
        tramaHexString = ''
        codificado = bytes.fromhex(tramaHexString)  # tramaHexString.encode()
        self.conectar.write(codificado + sparador)  # Señal para obtener datos
        print(f"MSG enviado: {codificado + sparador}")
        self.contador=0
        self.iniciarTemporizadorDescarga()

    def iniciarTemporizadorDescarga(self):
        self.total_pasos = 40 # Establece las veces que se repetira el timer
        self.AvanzePB=0.0
        self.DatosCompletos = b''
        self.primeraventana.progressBarDescarga.setValue(0)
        self.timer_intervalo = 10 # 200 ms
        self.timerDescarga.start(self.timer_intervalo)


    def actualizarContadorDescarga(self):

        self.contador += 1
        print(f"Contador: {self.contador}")

        DatosControlador = self.conectar.readline(255)
        print(f"DatosControlador: {DatosControlador}")
        self.DatosCompletos  = self.DatosCompletos  + DatosControlador
        self.primeraventana.progressBarDescarga.setValue(self.contador)

        valorAvanzePB = 100 / self.total_pasos
        self.AvanzePB = self.AvanzePB + valorAvanzePB
        self.primeraventana.progressBarDescarga.setValue(math.trunc(self.AvanzePB))
        if self.contador >= self.total_pasos:
            self.timerDescarga.stop()
            self.primeraventana.progressBarDescarga.setValue(100)
            self.conectar.close()
            print("DatosCompletos:")
            print(self.DatosCompletos )
            self.DatosCompletosGlobales=self.DatosCompletos

    def GenerarArchivoEXEL(self):
        cadena_bytes = self.DatosCompletosGlobales
        # Inicializar la lista para almacenar los resultados
        resultados = []

        # Iterar sobre la cadena de bytes
        i = 0
        while i < len(cadena_bytes):
            # Buscar 'C0x' o 'L0x'
            if cadena_bytes[i:i + 3] == b'C0x':
                num_bytes_siguientes = 20
                resultado = [
                    cadena_bytes[i:i + 1].decode(),  # L
                    cadena_bytes[i + 1:i + 7].decode(),  # 0x0000
                    float(cadena_bytes[i + 7:i + 11].decode()),  # 2.55
                    int(cadena_bytes[i + 11:i + 14].decode()),  # 003
                    ord(cadena_bytes[i + 14:i + 15]) - 90,  # Restar (-90)
                    cadena_bytes[i + 15:i + 23].decode()  # '1111111111' decodificado
                ]
                resultados.append(resultado)
                i += 3 + num_bytes_siguientes
            elif cadena_bytes[i:i + 3] == b'L0x':
                num_bytes_siguientes = 97
                resultado = [
                    cadena_bytes[i:i + 1].decode(),  # L
                    cadena_bytes[i + 1:i + 7].decode(),  # 0x0000
                    float(cadena_bytes[i + 7:i + 11].decode()),  # 2.55
                    int(cadena_bytes[i + 11:i + 14].decode()),  # 003
                    ord(cadena_bytes[i + 14:i + 15]) - 90,  # Restar (-90)
                    cadena_bytes[i + 15:i + 100].decode()  # '1111111111' decodificado
                ]
                resultados.append(resultado)
                i += 3 + num_bytes_siguientes
            else:
                i += 1

        print('Datos encontrados:')
        for resultado in resultados:
            print(resultado)
        # Crear un DataFrame con los resultados
        df = pd.DataFrame(resultados,
                          columns=['Tipo', 'No. Secuencia', 'Nivel Bateria[V]', 'Ptx[dBm]', 'Prx[dBm]', 'Carga Util'])
        print(df)
        # Crear una ventana de Tkinter (no se mostrará físicamente)
        # Crear una ventana de diálogo para seleccionar la ubicación y el nombre del archivo Excel
        root = Tk()
        root.withdraw()  # Ocultar la ventana principal de la aplicación

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        # Verificar si se proporcionó un nombre de archivo antes de guardar
        if file_path:
            # Guardar el DataFrame en el archivo Excel seleccionado
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Hoja1')

                # Ajustar automáticamente el tamaño de las celdas
                worksheet = writer.sheets['Hoja1']
                for column_cells in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column_cells]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                # Centrar el contenido en todas las celdas
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1,
                                               max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')

            print(f"DataFrame guardado exitosamente en {file_path}")
        else:
            print("Operación de guardado cancelada.")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ventana = mostrar()
    ventana.show()
    sys.exit(app.exec_())

'''     conectar.open()
        conectar.flush()
        deciInt = self.obtexto()
        deciStr = str(deciInt)
        MSG = 'ho' + "\r"
        conectar.write(MSG.encode())
        print('Se envió el MSG: ' + MSG)
        print('Se envió el MSG2 HEX: ')
        RET2 = conectar.read(3)
        print("LLEGADA:")
        print(RET2)
        conectar.close()'''

'''#RET1 = RET[:2].decode()# si el valor exadecimal es compatible con un digito ACII
RET1 = RET[:2].hex() # si el valor exadecimal no es compatible con un digito ACII
RET2 = RET[2:4].hex()
RET3 = RET[4:6].hex()
RET4 = RET[6:8].hex()
print(('0x' + RET1, RET2, RET3, RET4))
self.listaSubida.append(('0x'+RET1, '0x'+RET2, '0x'+RET3, '0x'+RET4))
RET = conectar.readline()
print("LLEGADA 1 bytes:")
print(RET)
#RET1 = RET[:2].decode()# si el valor exadecimal es compatible con un digito ACII
RET1 = RET[:2].hex() # si el valor exadecimal no es compatible con un digito ACII
RET2 = RET[2:4].hex()
RET3 = RET[4:6].hex()
RET4 = RET[6:8].hex()
print(('0x' + RET1, RET2, RET3, RET4))
self.listaSubida.append(('0x'+RET1, '0x'+RET2, '0x'+RET3, '0x'+RET4))
RET = conectar.readline()
print("LLEGADA 1 bytes:")
print(RET)
#RET1 = RET[:2].decode()# si el valor exadecimal es compatible con un digito ACII
RET1 = RET[:2].hex() # si el valor exadecimal no es compatible con un digito ACII
RET2 = RET[2:4].hex()
RET3 = RET[4:6].hex()
RET4 = RET[6:8].hex()
print(('0x' + RET1, RET2, RET3, RET4))
self.listaSubida.append(('0x'+RET1, '0x'+RET2, '0x'+RET3, '0x'+RET4))
'''

'''for x in range(50):
    m=y+5
    aux = RET2[y:m-1].decode()
    aux2 = RET2[m-1:m].hex()
    aux3 = int(aux2, 16)
    w = copy(open_workbook('datos_personales.xls'))

    if (str(aux3) != '0'):
        w.get_sheet(0).write(z, 1, '       ' + aux)
        w.get_sheet(0).write(z, 2, '        ' + str(aux3-90))
        #w.get_sheet(0).write(z, 3, '        ' + str(P))
        cons += 1
        w.get_sheet(0).write(cons + 1, 0, '        ' + str(cons))
        if (P == 0):
            w.get_sheet(0).write(z + 1, 0, '        ' + str(P))
        #auxNtx = auxNtx + Ntx
            # elif(str(aux3) == '0'):
            #    break
    else:

        w.get_sheet(0).write(z, 2, '        ')
        w.get_sheet(0).write(z, 3, '        ')
        w.get_sheet(0).write(z, 0, '        ')
        #z = z - 1

    #if (auxNtx - cons != 0):

    #elif(auxNtx - cons == 0):
        #w.get_sheet(0).write(auxNtx, 0, '' + str(np))
    #    auxNtx = auxNtx + Ntx
    w.save('datos_personales.xls')

    z += 1
    y = m'''

'''w = copy(open_workbook('datos_personales.xls'))

for i in range(Ntx):# LED
    G=H+i+2
    w.get_sheet(0).write(G, 3, '       ' + str(P))
    if (i % 2 == 0):
        self.primeraventana.progressBar_2.setValue(100)
        time.sleep(NT)
    else:
        self.primeraventana.progressBar_2.setValue(0)
        time.sleep(NT)
H=G-1
w.save('datos_personales.xls')'''
'''
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl.styles import Alignment

cadena_bytes = b'C0x00002.55003$11111111L0x00012.55003$1111111111C0x00022.55003$11111111'

# ... (código de procesamiento)

# Crear un DataFrame con los resultados
df = pd.DataFrame(resultados, columns=['Tipo', 'Código', 'Número', 'Valor', 'Resta', 'Cadena'])

# Crear una ventana de diálogo para seleccionar la ubicación y el nombre del archivo Excel
root = Tk()
root.withdraw()  # Ocultar la ventana principal de la aplicación

file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

# Verificar si se proporcionó un nombre de archivo antes de guardar
if file_path:
    # Guardar el DataFrame en el archivo Excel seleccionado
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hoja1')

        # Ajustar automáticamente el tamaño de las celdas
        worksheet = writer.sheets['Hoja1']
        for column_cells in worksheet.columns:
            max_length = 0
            column = [cell for cell in column_cells]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Centrar el contenido en todas las celdas
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"DataFrame guardado exitosamente en {file_path}")
else:
    print("Operación de guardado cancelada.")'''

# Guardar el DataFrame en el archivo seleccionado
# df.to_excel(file_path, index=False)
# print(f"Archivo guardado en: {file_path}")

################
# Guardar el DataFrame en un archivo Excel
# df.to_excel('resultados.xlsx', index=False)
# df.to_csv('resultados.csv', index=False)
'''
conteo_CL0x = 0
# Definir las subcadenas a buscar
subcadenas_a_buscar = [b'C0x', b'L0x']
# Iterar sobre la cadena de bytes
for subcadena_a_buscar in subcadenas_a_buscar:
    for i in range(len(DatosCompletos) - len(subcadena_a_buscar) + 1):
        subcadena_actual = DatosCompletos[i:i + len(subcadena_a_buscar)]
        if subcadena_actual == subcadena_a_buscar:
            conteo_CL0x += 1
# Mostrar el resultado total
print(f"Total de contador CL0x: {conteo_CL0x}")

        self.conectar.open()
        self.conectar.flush()
        sparador = b'\r'
        tramaHexString = ''
        codificado = bytes.fromhex(tramaHexString)  # tramaHexString.encode()
        self.conectar.write(codificado + sparador)  # Señal para obtener datos
        print('MSG HEX enviado: ')
        print(codificado + sparador)
        DatosCompletos = b''
        RangoLecturaMemoria = 40
        aumento = 100 / RangoLecturaMemoria
        ContPGBDes = 0.0

        #self.contador = 0
        #self.primeraventana.progressBarDescarga.setValue(0)
        for i in range(RangoLecturaMemoria):
            DatosControlador = self.conectar.readline(255)
            ContPGBDes = ContPGBDes + aumento
            self.primeraventana.progressBarDescarga.setValue(math.trunc(ContPGBDes))
            print('DatosControlador')
            print(DatosControlador)
            DatosCompletos = DatosCompletos + DatosControlador
        self.conectar.close()
        print(DatosCompletos)

'''
